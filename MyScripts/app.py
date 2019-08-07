import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']#selecting sheets
#cell = sheet['a1']#selecting cells
#cell = sheet.cell(1,1)#selecting cells version 2

#how many rows do we have in the spreadsheet
print(sheet.max_row)#getting last row
for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * .9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price
#openpyxl module method to grab values from a sheet via selecting the sheet and then the range of rows and columns as seen below
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)
#Create a BarChart object
chart = BarChart()
#add the values data to it
chart.add_data(values)
#place it in the spreadsheet at the specified cell
sheet.add_chart(chart,'E2')
wb.save('transactions2.xlsx')